from openpyxl.reader.excel import load_workbook

import datetime

from db_def import db
from db_def import Account
from db_def import Note
from db_def import Media
from db_def import Context
from db_def import Feedback
from db_def import Site

db.drop_all()
db.create_all()

deployment = False


wb = load_workbook(filename = r'data.xlsx')
context_sheet = wb.get_sheet_by_name(name = 'Context')
site_sheet = wb.get_sheet_by_name(name = 'Site')
account_sheet = wb.get_sheet_by_name(name = 'Account')
note_sheet = wb.get_sheet_by_name(name = 'Note')
feedback_sheet = wb.get_sheet_by_name(name = 'Feedback')

n = site_sheet.cell('A1').value
for i in range(2,2+n):
	name = site_sheet.cell('B' + str(i)).value	
	description  = site_sheet.cell('C' + str(i)).value
	image_url  = site_sheet.cell('D' + str(i)).value

	site = Site(name, description)
	site.image_url = image_url
	db.session.add(site)
	print "create site: %s" % site		
db.session.commit()		

created_at = datetime.date(2014,3,1)

if not deployment:
	n = account_sheet.cell('A1').value
	for i in range(2,2+n):
		id = account_sheet.cell('A' + str(i)).value	
		username = account_sheet.cell('B' + str(i)).value
		account = Account(username)
		account.name = account_sheet.cell('C' + str(i)).value
		account.email = account_sheet.cell('D' + str(i)).value
		account.password = account_sheet.cell('E' + str(i)).value
		account.consent = account_sheet.cell('F' + str(i)).value
		account.icon_url = account_sheet.cell('G' + str(i)).value

		created_at += datetime.timedelta(days=1)
		account.created_at = created_at
		account.modified_at =  created_at
		if id:
			print "create account: %s" % account
			db.session.add(account)
	db.session.commit()
else:
	account = Account("default")
	created_at += datetime.timedelta(days=1)
	account.created_at = created_at
	account.modified_at =  created_at
	db.session.add(account)
	db.session.commit()

n = context_sheet.cell('A1').value
for i in range(2,2+n):
	id = note_sheet.cell('A' + str(i)).value	
	kind = context_sheet.cell('B' + str(i)).value
	name = context_sheet.cell('C' + str(i)).value
	title = context_sheet.cell('D' + str(i)).value
	description = context_sheet.cell('E' + str(i)).value
	site = context_sheet.cell('F' + str(i)).value
	# extras = context_sheet.cell('G' + str(i)).value
	context = Context(kind, name, title, description)
	# context.extras = extras
	
	site = Site.query.filter_by(name=site).first()
	context.site_id = site.id

	if kind == 'Landmark':
		lat = context_sheet.cell('G' + str(i)).value
		lgn = context_sheet.cell('H' + str(i)).value
		context.extras = str({"latitude": float(lat), "longitude": float(lgn)})

	if id:
		print "create context: %s" % context
		db.session.add(context)
db.session.commit()

created_at = datetime.datetime(2014,4,1,0,0)


from random import randint

if not deployment:
	# NOTE
	n = note_sheet.cell('A1').value
	for i in range(2,2+n):
		id = note_sheet.cell('A' + str(i)).value	
		username = note_sheet.cell('B' + str(i)).value	
		context  = note_sheet.cell('C' + str(i)).value
		kind     = note_sheet.cell('D' + str(i)).value
		content  = note_sheet.cell('E' + str(i)).value	

		media_kind  = note_sheet.cell('F' + str(i)).value	
		media_title = note_sheet.cell('G' + str(i)).value	
		media_url = note_sheet.cell('H' + str(i)).value	
		# media_url = "https://dl.dropboxusercontent.com/u/5104407/nntest/" + media_url


		latitude  = note_sheet.cell('I' + str(i)).value	
		longitude  = note_sheet.cell('J' + str(i)).value		
		created_at = note_sheet.cell('K' + str(i)).value

		if not created_at:
			created_at = 1396325280
			det1 = 1.0 + float(randint(1,100) - 50)/5000000
			det2 = 1.0 + float(randint(1,100) - 50)/5000000				
		else:
			created_at = int(created_at)
			det1 = 1.0
			det2 = 1.0

		date = datetime.datetime.fromtimestamp(created_at)	
		if id:
			a = Account.query.filter_by(username=username).first()
			c = Context.query.filter_by(name=context).first()
			note = Note(a.id, c.id, kind, content)

			note.latitude = float(latitude) * det1
			note.longitude = float(longitude) * det2		
			note.created_at =  date
			note.modified_at =  date
			# created_at += datetime.timedelta(seconds=100)		
			print "create note: %s" % note
			db.session.add(note)
			db.session.commit()
			
			if media_kind:
				media = Media(note.id, media_kind, media_title, media_url) 
				media.created_at = date
				# created_at += datetime.timedelta(seconds=60)		
				print "create media: %s" % media
				db.session.add(media)
				db.session.commit()

	n = feedback_sheet.cell('A1').value
	for i in range(2,2+n):
		table_name = feedback_sheet.cell('B' + str(i)).value	
		row_id  = feedback_sheet.cell('C' + str(i)).value
		kind     = feedback_sheet.cell('D' + str(i)).value
		content  = feedback_sheet.cell('E' + str(i)).value	
		username = feedback_sheet.cell('F' + str(i)).value		

		if id:
			a = Account.query.filter_by(username=username).first()
			feedback = Feedback(a.id, kind, content, table_name, row_id)
			db.session.add(feedback)
			
			print "create feedback: %s" % feedback		
	db.session.commit()		


#print sheet_ranges.cell('B2').value # D18
